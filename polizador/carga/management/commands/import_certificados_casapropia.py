import calendar
import os
import re
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from simple_history.utils import bulk_create_with_history

from carga.models import Certificado, Obra

_DOS_DECIMALES = Decimal("0.01")
_TRES_DECIMALES = Decimal("0.001")
_ENTERO = Decimal("1")

DEFAULT_XLSX = os.path.join(settings.BASE_DIR, "..", "env", ".snipets", "casapropia.xlsx")
DEFAULT_DUMP = os.path.join(settings.BASE_DIR, "dump.txt")
HOJA = "CASA PROPIA"


def _decimal(valor, exp):
    if valor is None or valor == "":
        return Decimal(0).quantize(exp)
    try:
        return Decimal(str(valor)).quantize(exp, rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return Decimal(0).quantize(exp)


def _extraer_numero(valor):
    """certificado_rubro_obra/anticipo/devanticipo a veces llegan como texto libre
    (ej. "4 COMPL.", "REAJUSTE", "DESEMBOLSO 3") en vez de un número entero: se toma
    el primer entero encontrado en el texto, o 0 si no hay ninguno."""
    if valor is None or valor == "":
        return Decimal(0)
    if isinstance(valor, (int, float, Decimal)):
        return _decimal(valor, _ENTERO)
    match = re.search(r"\d+", str(valor))
    return Decimal(match.group()) if match else Decimal(0)


def _financiamiento(programa_texto):
    """Las obras cuyo programa (columna "PROGRAMA " del excel) menciona FO.PRO.VI. son
    financiamiento Provincia; el resto (Programa Casa Propia - Construir Futuro
    "clásico") es Nación."""
    if programa_texto and "fo.pro.vi" in programa_texto.lower():
        return "P"
    return "N"


def _parse_fecha(valor, periodo):
    """certificado_fecha viene como texto dd/mm/aaaa. Cuando falta (72 filas del
    excel), se usa el último día del mes de "periodo" como aproximación -- coincide
    con el patrón del resto de las filas, donde certificado_fecha siempre cae en el
    último día del mes de periodo."""
    if isinstance(valor, str) and valor.strip():
        try:
            return datetime.strptime(valor.strip(), "%d/%m/%Y").date()
        except ValueError:
            pass
    if isinstance(periodo, datetime):
        ultimo_dia = calendar.monthrange(periodo.year, periodo.month)[1]
        return date(periodo.year, periodo.month, ultimo_dia)
    return None


class Command(BaseCommand):
    """
    Management command: importa certificados históricos del Programa Casa Propia -
    Construir Futuro desde env/.snipets/casapropia.xlsx (hoja "CASA PROPIA") a
    carga.models.Certificado.

    El match de cada fila con la Obra existente se hace por obra_convenio (comparación
    case-insensitive por prefijo: el excel trae el convenio truncado respecto al
    guardado en la base, ej. "CONVE-2021-124071226" vs
    "CONVE-2021-124071226-APN-DGDYD#MDTYH"). Si no matchea por convenio, se intenta un
    fallback por obra_expediente exacto (cubre un puñado de erratas puntuales del
    convenio en el excel, verificadas a mano contra la base).

    Los certificados se cargan con certificado_tipo="LEGACY" y
    certificado_fecha_carga_legacy=True, replicando la regla de
    CrearCertificado.form_valid (carga/views/certificadoviews.py): para certificados
    legacy, certificado_fecha_carga = certificado_fecha en vez de la fecha real de
    importación. certificado_rubro_db es siempre Vivienda. certificado_financiamiento
    es Provincia si la columna "PROGRAMA " de la fila menciona FO.PRO.VI. (el programa
    fue renombrado a mitad del período cubierto por el excel), y Nación en caso
    contrario.

    Es re-ejecutable: antes de crear cada certificado se chequea que su
    certificado_expediente no exista ya en carga.Certificado (en ningún certificado,
    no sólo legacy) -- si ya existe, la fila se omite sin tocarlo. Esto también
    absorbe expedientes repetidos dentro del propio excel (sólo se crea el primero).
    No aplica a filas sin certificado_expediente (quedan fuera de este filtro).

    Corre en modo dry-run por default (no escribe nada); pasar --commit para persistir.

    Uso:
        python manage.py import_certificados_casapropia
        python manage.py import_certificados_casapropia --commit
        python manage.py import_certificados_casapropia --file /ruta/a/otro.xlsx --commit
    """

    help = "Importa certificados históricos de Casa Propia desde casapropia.xlsx (dry-run por default, pasar --commit para persistir)."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_XLSX, help="Ruta al archivo .xlsx de origen.")
        parser.add_argument("--commit", action="store_true", help="Persiste los cambios (sin esto, sólo reporta).")
        parser.add_argument("--dump-file", default=DEFAULT_DUMP, help="Ruta del archivo donde se vuelcan las filas con problemas, para revisión manual.")

    def handle(self, *args, **options):
        filepath = os.path.abspath(options["file"])
        commit = options["commit"]

        if not os.path.isfile(filepath):
            self.stderr.write(self.style.ERROR(f"No existe el archivo: {filepath}"))
            return

        wb = openpyxl.load_workbook(filepath, data_only=True)
        if HOJA not in wb.sheetnames:
            self.stderr.write(self.style.ERROR(f"El archivo no tiene una hoja '{HOJA}'"))
            return
        ws = wb[HOJA]

        filas = list(ws.iter_rows(min_row=1, values_only=True))
        header = filas[0]
        registros = [dict(zip(header, fila)) for fila in filas[1:] if any(c is not None for c in fila)]
        self.stdout.write(f"Filas con datos en el excel: {len(registros)}")

        obras_por_convenio = list(
            Obra.objects.exclude(obra_convenio__isnull=True)
            .exclude(obra_convenio="")
            .values_list("id", "obra_convenio")
        )
        obras_por_expediente = {}
        for id_, expte in Obra.objects.exclude(obra_expediente="").values_list("id", "obra_expediente"):
            obras_por_expediente.setdefault(expte.strip(), []).append(id_)

        # Primer filtro: si certificado_expediente ya existe en carga.Certificado (en
        # cualquier certificado, no sólo legacy), no se toca -- se asume ya cargado.
        expedientes_existentes = set(
            Certificado.objects.exclude(certificado_expediente="").values_list(
                "certificado_expediente", flat=True
            )
        )

        sin_match = []
        sin_fecha = []
        rubro_no_numerico = []
        expediente_truncado = []
        ya_existentes = 0
        a_crear = []

        for r in registros:
            convenio = (r.get("obra_convenio") or "").strip()
            obra_nombre = (r.get("obra_nombre") or "").strip()
            empresa = (r.get("obra_empresa") or "").strip()
            identificacion = (convenio, obra_nombre, empresa)

            expediente_original = (r.get("certificado_expediente") or "").strip()
            expediente = expediente_original[:18] if len(expediente_original) > 18 else expediente_original

            if expediente and expediente in expedientes_existentes:
                ya_existentes += 1
                continue

            if expediente_original != expediente:
                expediente_truncado.append((*identificacion, f"certificado_expediente truncado a 18 caracteres: {expediente_original!r} -> {expediente!r}"))

            expediente_obra = (r.get("obra_expediente") or "").strip()

            obra_id = None
            if convenio:
                matches = [id_ for id_, c in obras_por_convenio if c.upper().startswith(convenio.upper())]
                if len(matches) == 1:
                    obra_id = matches[0]
            if obra_id is None and expediente_obra:
                candidatos = obras_por_expediente.get(expediente_obra, [])
                if len(candidatos) == 1:
                    obra_id = candidatos[0]

            if obra_id is None:
                sin_match.append((*identificacion, f"no se encontró Obra (obra_expediente={expediente_obra!r})"))
                continue

            periodo = r.get("periodo")
            fecha = _parse_fecha(r.get("certificado_fecha"), periodo)
            if fecha is None:
                sin_fecha.append((*identificacion, f"sin fecha determinable (certificado_expediente={expediente!r})"))
                continue

            for campo in ("certificado_rubro_obra", "certificado_rubro_anticipo", "certificado_rubro_devanticipo"):
                valor = r.get(campo)
                if valor is not None and not isinstance(valor, (int, float)):
                    rubro_no_numerico.append((*identificacion, f"{campo}={valor!r} no es numérico, se usó el primer entero encontrado"))

            monto_pesos = _decimal(r.get("certificado_monto_pesos"), _DOS_DECIMALES)
            rubro_obra = _extraer_numero(r.get("certificado_rubro_obra"))
            rubro_anticipo = _extraer_numero(r.get("certificado_rubro_anticipo"))
            rubro_devanticipo = _extraer_numero(r.get("certificado_rubro_devanticipo"))

            if expediente:
                expedientes_existentes.add(expediente)

            periodo_str = periodo.strftime("%m/%Y") if isinstance(periodo, datetime) else None

            a_crear.append(Certificado(
                certificado_obra_id=obra_id,
                certificado_tipo="LEGACY",
                certificado_financiamiento=_financiamiento(r.get("PROGRAMA ")),
                certificado_rubro_db_id=1,  # Vivienda
                certificado_expediente=expediente,
                certificado_periodo=periodo_str,
                certificado_fecha=fecha,
                certificado_fecha_carga=fecha,
                certificado_fecha_carga_legacy=True,
                certificado_rubro_anticipo=rubro_anticipo,
                certificado_rubro_obra=rubro_obra,
                certificado_rubro_devanticipo=rubro_devanticipo,
                certificado_monto_pesos=monto_pesos,
                certificado_monto_uvi=_decimal(r.get("certificado_monto_uvi"), _DOS_DECIMALES),
                certificado_mes_pct=_decimal(r.get("certificado_mes_pct"), _TRES_DECIMALES),
                certificado_ante_pct=_decimal(r.get("certificado_ante_pct"), _TRES_DECIMALES),
                certificado_acum_pct=_decimal(r.get("certificado_acum_pct"), _TRES_DECIMALES),
                certificado_descuento_anticipo_pesos=_decimal(r.get("certificado_descuento_anticipo_pesos"), _DOS_DECIMALES),
                certificado_descuento_anticipo_uvi=_decimal(r.get("certificado_descuento_anticipo_uvi"), _DOS_DECIMALES),
            ))

        self._reportar(
            sin_match, sin_fecha, rubro_no_numerico, expediente_truncado, ya_existentes, a_crear,
            options["dump_file"],
        )

        if not a_crear:
            return

        if not commit:
            self.stdout.write(self.style.WARNING(
                f"Dry-run: se crearían {len(a_crear)} certificados. Volver a correr con --commit para persistir."
            ))
            return

        with transaction.atomic():
            bulk_create_with_history(a_crear, Certificado, batch_size=500)
        self.stdout.write(self.style.SUCCESS(f"Se crearon {len(a_crear)} certificados."))

    def _reportar(self, sin_match, sin_fecha, rubro_no_numerico, expediente_truncado, ya_existentes, a_crear, dump_file):
        categorias = (
            ("Filas sin Obra matcheada", sin_match),
            ("Filas sin fecha determinable", sin_fecha),
            ("Filas con número de rubro no numérico (se extrajo el primer entero encontrado)", rubro_no_numerico),
            ("Filas con certificado_expediente > 18 caracteres (se truncó)", expediente_truncado),
        )

        for titulo, entradas in categorias:
            if entradas:
                self.stdout.write(self.style.WARNING(f"{titulo}: {len(entradas)} (detalle en {dump_file})"))

        if ya_existentes:
            self.stdout.write(
                f"Filas cuyo certificado_expediente ya existe en carga.Certificado, omitidas: {ya_existentes}"
            )

        self.stdout.write(f"Certificados a crear: {len(a_crear)}")

        if any(entradas for _, entradas in categorias):
            with open(dump_file, "w", encoding="utf-8") as f:
                for titulo, entradas in categorias:
                    if not entradas:
                        continue
                    f.write(f"=== {titulo} ({len(entradas)}) ===\n")
                    for convenio, obra_nombre, empresa, detalle in entradas:
                        f.write(f"convenio={convenio!r} | obra={obra_nombre!r} | empresa={empresa!r} | {detalle}\n")
                    f.write("\n")
            self.stdout.write(f"Detalle de filas con problemas volcado en: {dump_file}")
