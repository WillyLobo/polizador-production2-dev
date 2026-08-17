import logging
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from personalizador.management.commands.importar_informe_ipduv import _blank, _normalize_text
from personalizador.licencias import LICENCIA_IMPORTADA_MOTIVO_PREFIJO as TAG
from personalizador.models import Agente, LicenciaPermiso, TipoLicenciaPermiso

logger = logging.getLogger(__name__)

DEFAULT_PATH = settings.BASE_DIR.parent / "env" / ".snipets" / "control_licencias.xlsx"

# TAG (importado de personalizador.licencias) es el prefijo fijo de
# licenciapermiso_motivo para los registros generados por este comando:
# permite reconocerlos en una re-ejecucion (buscar por agente+tipo+
# fecha_desde+motivo__startswith=TAG) y actualizarlos en vez de duplicarlos,
# sin tocar licencias cargadas a mano para la misma fecha. El mismo prefijo
# se usa en api/views/personalizador_views.py para la columna/filtro
# "Importada" del datatable.

# "ART 50 Inc 4" (Razones Particulares con compensacion horaria) esta cargado
# en el excel en dias equivalentes, pero el tipo en la base esta en horas
# (unidad HS, tope 40/anio): se convierte multiplicando por la jornada
# laboral estandar del organismo (6:30 hs), redondeando al entero mas cercano.
HORAS_POR_JORNADA = 6.5

# Los "tipos" de columna de licencia que este comando entiende, mapeados a la
# clave natural (categoria, nombre) de TipoLicenciaPermiso. "art10" se procesa
# aparte (segunda pasada, ver handle()) porque su validacion en el modelo
# depende de que la "Anual" del anio siguiente ya este cargada.
TIPO_KIND = {
    "anual": ("LOR", "Anual"),
    "art10": ("LOR", "Anual Proporcional"),
    "invierno": ("LOR", "Anual de Invierno"),
    "art50_4": ("PER", "Razones Particulares"),
    "art50_11": ("PER", "Atención a Familiar Incapacitado o Discapacitado"),
    "natalicio": ("PER", "Natalicio del Agente"),
    "matrimonio": ("LEX", "Matrimonio"),
}

# Columnas 0-indexadas de cada hoja "L.O.A. <año>", verificadas contra el
# archivo real. El orden y la cantidad de columnas difieren entre hojas
# (el excel se fue modificando año a año), asi que se hardcodean por hoja en
# vez de intentar leerlas por rotulo (que ademas es inconsistente: algunas
# hojas dicen "APELLIDO Y NOMBRE" y otras "NOMBRE Y APELLIDO" para el mismo
# formato real "APELLIDO NOMBRE(S)").
#
# Cada hoja tiene, ademas de "anual"/"invierno"/etc., una columna de Art. 10
# rotulada con el año de la hoja (col_art10_recap) que es un recordatorio del
# adelanto ya tomado el año anterior (mismo dato que la col_art10_nuevo de la
# hoja anterior) y NO se importa para no duplicar; y una columna rotulada con
# año+1 (dentro de "columnas", kind="art10") que es el adelanto nuevo de ese
# año contra el cupo del año siguiente, y si se importa.
LOA_SHEETS = [
    {
        "sheet": "L.O.A. 2025",
        "anio": 2025,
        "col_nombre": 1,
        "col_devoluciones": 9,
        "col_art10_recap": 5,
        "columnas": [
            (4, "anual"),
            (6, "art10"),
            (8, "art50_4"),
            (10, "invierno"),
            (11, "art50_11"),
            (14, "natalicio"),
        ],
    },
    {
        "sheet": "L.O.A. 2024",
        "anio": 2024,
        "col_nombre": 0,
        "col_devoluciones": 8,
        "col_art10_recap": 4,
        "columnas": [
            (3, "anual"),
            (5, "art10"),
            (7, "art50_4"),
            (9, "invierno"),
            (10, "art50_11"),
        ],
    },
    {
        # Ojo: el nombre real de la hoja tiene un espacio al principio y no
        # tiene puntos ("L.O.A" en vez de "L.O.A."), a diferencia de las otras
        # dos hojas de años. wb.sheetnames lo confirma tal cual esta aca.
        "sheet": " L.O.A 2023",
        "anio": 2023,
        "col_nombre": 0,
        "col_devoluciones": 7,
        "col_art10_recap": None,  # esta hoja no tiene columna de recap, solo la nueva.
        "columnas": [
            (3, "anual"),
            (4, "art10"),
            (6, "art50_4"),
            (8, "invierno"),
            (9, "art50_11"),
            (10, "matrimonio"),
        ],
    },
]
LOA_FIRST_DATA_ROW = 2

# Las 3 hojas de Motivos de Salud comparten exactamente la misma disposicion
# de columnas (0-indexadas); solo cambia el nombre de la hoja y el año que
# representa. La columna 7 (Maternidad) no tiene rotulo en "MOTIVO DE SALUD
# 24" pero la posicion es la misma que en las otras dos hojas.
SALUD_SHEETS = [
    {"sheet": "MOTIVOS DE SALUD 2025", "anio": 2025},
    {"sheet": "MOTIVOS DE SALUD 2026", "anio": 2026},
    {"sheet": "MOTIVO DE SALUD 24", "anio": 2024},
]
SALUD_COL_NOMBRE = 0
SALUD_COLUMNAS = [
    (2, "LOR", "Motivos de Salud (Enfermedad o lesión común)"),
    (3, "LOR", "Motivos de Salud (Enfermedad prolongada o lesión grave)"),
    (4, "LOR", "Motivos de Salud (Accidente de trabajo o enfermedad ocupacional)"),
    (5, "LOR", "Motivos de Salud (Atención de familiar enfermo)"),
    (6, "LOR", "Motivos de Salud (Enfermedad o lesión grave de familiar)"),
    (7, "LOR", "Maternidad o Gravidez"),
]
SALUD_FIRST_DATA_ROW = 3
SALUD_HEADER_ROW = 2


def _tokens_nombre(raw):
    """Normaliza un nombre completo a un conjunto de palabras (mayusculas,
    sin comas/puntos, sin tildes, sin espacios repetidos) para matchear
    contra Agente sin depender del orden apellido/nombre (inconsistente
    entre hojas) ni de comas pegadas al apellido (ej. 'TORRES,MIGUEL')."""
    texto = _normalize_text(raw).upper().replace(",", " ").replace(".", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return frozenset(t for t in texto.split() if t)


def _parse_numero(raw):
    if isinstance(raw, (int, float)):
        return float(raw), None
    texto = str(raw).strip().replace(",", ".")
    try:
        return float(texto), None
    except ValueError:
        return None, f"valor no numerico ({raw!r})"


def _cantidad(kind, valor):
    if kind == "art50_4":
        return round(valor * HORAS_POR_JORNADA)
    return round(valor)


def _registrar_aviso(avisos, hoja, fila, agente_desc, cols, msg, silencioso=False):
    """Guarda un aviso con la hoja/fila/columnas del xlsx a las que se
    refiere, para poder resaltarlas despues en el xlsx de errores (ver
    _exportar_errores). Si silencioso=True, el aviso va igual al xlsx pero no
    se imprime por consola (ver el filtro en handle())."""
    if cols is None:
        cols = []
    elif isinstance(cols, int):
        cols = [cols]
    avisos.append({"hoja": hoja, "fila": fila, "agente_desc": agente_desc, "cols": list(cols), "msg": msg, "silencioso": silencioso})


class Command(BaseCommand):
    help = (
        "Importa control_licencias.xlsx: 3 hojas 'L.O.A. <año>' (Licencia Anual "
        "Ordinaria/Proporcional/de Invierno, Art. 50 inc. 4/11, Natalicio, "
        "Matrimonio) y 3 hojas de Motivos de Salud (Art. 14 A-E + Maternidad). "
        "Cada fila trae un TOTAL ANUAL usado por agente y tipo, sin fecha "
        "puntual, asi que por cada celda con datos se crea/actualiza UN "
        "LicenciaPermiso sintetico con fecha_desde=1/1 del año que "
        "corresponde (año de la hoja, o año+1 para 'Anual de Invierno' y "
        "para el adelanto Art. 10, que ya aplica su propio corrimiento de "
        "año en el modelo) y licenciapermiso_motivo empezando con "
        f"'{TAG}'. Ese prefijo es lo que permite re-correr el import sin "
        "duplicar (busca un registro existente por agente+tipo+fecha_desde+"
        "ese prefijo y lo actualiza en vez de crear uno nuevo) sin pisar "
        "licencias cargadas a mano para la misma fecha. "
        "El excel no trae DNI: cada fila matchea contra personalizador.Agente "
        "por nombre completo normalizado (sin tildes/mayusculas/orden), y si "
        "no hay exactamente un Agente candidato la fila se saltea y queda "
        "en el xlsx de errores. La columna DEVOLUCIONES (valores mezclados: "
        "numeros positivos/negativos y texto libre como 'D3') no se importa, "
        "solo se lista en el xlsx de errores para revision manual. Cada "
        "registro pasa por LicenciaPermiso.full_clean() antes de guardarse; "
        "si el modelo lo rechaza (ej. supera el cupo disponible del Art. 10) "
        "se saltea y queda como aviso."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", default=str(DEFAULT_PATH),
            help="Ruta al xlsx a importar (default: %(default)s)",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="No guarda cambios en la base de datos, solo informa.",
        )
        parser.add_argument(
            "--errores-file", default=None,
            help="Ruta donde guardar el xlsx con las filas que tuvieron avisos "
                 "(default: el archivo de entrada con sufijo '_errores').",
        )

    def handle(self, *args, **options):
        import openpyxl

        path = options["file"]
        dry_run = options["dry_run"]
        verbosity = options.get("verbosity", 1)

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontro el archivo: {path}")

        for cfg in LOA_SHEETS + SALUD_SHEETS:
            if cfg["sheet"] not in wb.sheetnames:
                raise CommandError(f"No se encontro la hoja '{cfg['sheet']}' en {path}.")

        tipos_cache = {}

        def get_tipo(categoria, nombre):
            key = (categoria, nombre)
            if key not in tipos_cache:
                tipos_cache[key] = TipoLicenciaPermiso.objects.filter(
                    tipolicenciapermiso_categoria=categoria, tipolicenciapermiso_nombre=nombre,
                ).first()
            return tipos_cache[key]

        self._cache_candidatos = {}
        indice_agentes = self._construir_indice_agentes()

        stats = defaultdict(int)
        avisos = []
        filas_vistas = {}
        headers_por_hoja = {}
        pend_normales = []
        pend_art10 = []

        with transaction.atomic():
            for cfg in LOA_SHEETS:
                self._procesar_hoja_loa(
                    wb, cfg, indice_agentes, get_tipo, avisos, filas_vistas,
                    headers_por_hoja, stats, pend_normales, pend_art10,
                )
            for cfg in SALUD_SHEETS:
                self._procesar_hoja_salud(
                    wb, cfg, indice_agentes, get_tipo, avisos, filas_vistas,
                    headers_por_hoja, stats, pend_normales,
                )

            for item in pend_normales + pend_art10:
                self._guardar_licencia(item, avisos, stats)

            if dry_run:
                transaction.set_rollback(True)

        if verbosity >= 1:
            self.stdout.write("")
            if dry_run:
                self.stdout.write(self.style.NOTICE("Modo dry-run: no se guardo ningun cambio en la base de datos."))
            self.stdout.write(self.style.MIGRATE_HEADING("Resumen:"))
            self.stdout.write(self.style.SUCCESS(f"    LicenciaPermiso creadas: {stats['creados']}"))
            self.stdout.write(self.style.SUCCESS(f"    LicenciaPermiso actualizadas: {stats['actualizados']}"))
            self.stdout.write(f"    LicenciaPermiso sin cambios: {stats['sin_cambios']}")
            self.stdout.write(self.style.WARNING(f"    Filas sin Agente matcheado (0 o >1 candidatos): {stats['agente_no_encontrado']}"))
            self.stdout.write(self.style.WARNING(f"    Registros rechazados por validacion del modelo: {stats['rechazados']}"))
            self.stdout.write(self.style.WARNING(f"    Celdas de DEVOLUCIONES no importadas (revisar a mano): {stats['devoluciones_sin_importar']}"))

        avisos_a_mostrar = [a for a in avisos if not a["silencioso"]]
        avisos_silenciados = len(avisos) - len(avisos_a_mostrar)
        if avisos and verbosity >= 1:
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(
                f"Avisos ({len(avisos_a_mostrar)}"
                + (f", +{avisos_silenciados} silenciados -DEVOLUCIONES-, ver el xlsx de errores)" if avisos_silenciados else "")
                + "):"
            ))
            for aviso in avisos_a_mostrar:
                self.stdout.write(f"    {self.style.WARNING(aviso['msg'])}")

        if avisos:
            ruta_errores = options["errores_file"]
            if not ruta_errores:
                origen = Path(path)
                ruta_errores = str(origen.with_name(f"{origen.stem}_errores.xlsx"))
            self._exportar_errores(headers_por_hoja, filas_vistas, avisos, ruta_errores)
            if verbosity >= 1:
                self.stdout.write(self.style.NOTICE(f"Filas con avisos exportadas a: {ruta_errores}"))

    def _construir_indice_agentes(self):
        return [
            (_tokens_nombre(f"{agente.agente_apellidos} {agente.agente_nombres}"), agente)
            for agente in Agente.objects.all()
        ]

    def _candidatos(self, nombre_raw, indice):
        """Un Agente es candidato si TODAS las palabras del nombre del excel
        aparecen entre las palabras de su nombre completo (el excel suele
        traer solo apellido + primer nombre, sin segundo nombre; la base
        tiene el nombre completo). Cacheado por nombre normalizado: el mismo
        agente aparece en varias hojas/años."""
        clave = _tokens_nombre(nombre_raw)
        if clave not in self._cache_candidatos:
            self._cache_candidatos[clave] = (
                [agente for tokens, agente in indice if clave and clave.issubset(tokens)] if clave else []
            )
        return self._cache_candidatos[clave]

    def _resolver_agente(self, nombre_raw, indice, avisos, hoja, fila, col):
        candidatos = self._candidatos(nombre_raw, indice)
        if len(candidatos) == 1:
            return candidatos[0]
        if not candidatos:
            _registrar_aviso(
                avisos, hoja, fila, nombre_raw, col,
                f"'{nombre_raw}': no se encontro ningun Agente con ese nombre, fila salteada.",
            )
        else:
            _registrar_aviso(
                avisos, hoja, fila, nombre_raw, col,
                f"'{nombre_raw}': coincide con {len(candidatos)} Agentes distintos, fila salteada (revisar a mano).",
            )
        return None

    def _procesar_hoja_loa(self, wb, cfg, indice, get_tipo, avisos, filas_vistas, headers_por_hoja, stats, pend_normales, pend_art10):
        ws = wb[cfg["sheet"]]
        headers_por_hoja[cfg["sheet"]] = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        for fila_excel, row in enumerate(
            ws.iter_rows(min_row=LOA_FIRST_DATA_ROW, values_only=True), start=LOA_FIRST_DATA_ROW,
        ):
            nombre_raw = row[cfg["col_nombre"]]
            if _blank(nombre_raw):
                continue
            filas_vistas[(cfg["sheet"], fila_excel)] = row

            if not _blank(row[cfg["col_devoluciones"]]):
                _registrar_aviso(
                    avisos, cfg["sheet"], fila_excel, nombre_raw, cfg["col_devoluciones"],
                    f"'{nombre_raw}': DEVOLUCIONES={row[cfg['col_devoluciones']]!r}, no se importa (revisar a mano).",
                    silencioso=True,
                )
                stats["devoluciones_sin_importar"] += 1

            agente = self._resolver_agente(nombre_raw, indice, avisos, cfg["sheet"], fila_excel, cfg["col_nombre"])
            if agente is None:
                stats["agente_no_encontrado"] += 1
                continue

            for col, kind in cfg["columnas"]:
                raw = row[col]
                if _blank(raw):
                    continue

                valor, error = _parse_numero(raw)
                categoria, nombre_tipo = TIPO_KIND[kind]
                if error:
                    _registrar_aviso(
                        avisos, cfg["sheet"], fila_excel, nombre_raw, col,
                        f"'{nombre_raw}': {nombre_tipo}: {error}, se salteo la columna.",
                    )
                    continue

                tipo = get_tipo(categoria, nombre_tipo)
                if tipo is None:
                    _registrar_aviso(
                        avisos, cfg["sheet"], fila_excel, nombre_raw, col,
                        f"'{nombre_raw}': no existe TipoLicenciaPermiso '{nombre_tipo}' ({categoria}) en el catalogo, se salteo la columna.",
                    )
                    continue

                anio = cfg["anio"] + (1 if kind == "invierno" else 0)
                item = {
                    "agente": agente, "agente_desc": nombre_raw, "tipo": tipo,
                    "fecha": date(anio, 1, 1), "cantidad": _cantidad(kind, valor),
                    "hoja": cfg["sheet"], "anio_hoja": cfg["anio"], "fila": fila_excel, "col": col,
                }
                (pend_art10 if kind == "art10" else pend_normales).append(item)

    def _procesar_hoja_salud(self, wb, cfg, indice, get_tipo, avisos, filas_vistas, headers_por_hoja, stats, pend_normales):
        ws = wb[cfg["sheet"]]
        headers_por_hoja[cfg["sheet"]] = [ws.cell(row=SALUD_HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]

        for fila_excel, row in enumerate(
            ws.iter_rows(min_row=SALUD_FIRST_DATA_ROW, values_only=True), start=SALUD_FIRST_DATA_ROW,
        ):
            nombre_raw = row[SALUD_COL_NOMBRE]
            if _blank(nombre_raw):
                continue
            filas_vistas[(cfg["sheet"], fila_excel)] = row

            agente = self._resolver_agente(nombre_raw, indice, avisos, cfg["sheet"], fila_excel, SALUD_COL_NOMBRE)
            if agente is None:
                stats["agente_no_encontrado"] += 1
                continue

            for col, categoria, nombre_tipo in SALUD_COLUMNAS:
                raw = row[col]
                if _blank(raw):
                    continue

                valor, error = _parse_numero(raw)
                if error:
                    _registrar_aviso(
                        avisos, cfg["sheet"], fila_excel, nombre_raw, col,
                        f"'{nombre_raw}': {nombre_tipo}: {error}, se salteo la columna.",
                    )
                    continue

                tipo = get_tipo(categoria, nombre_tipo)
                if tipo is None:
                    _registrar_aviso(
                        avisos, cfg["sheet"], fila_excel, nombre_raw, col,
                        f"'{nombre_raw}': no existe TipoLicenciaPermiso '{nombre_tipo}' ({categoria}) en el catalogo, se salteo la columna.",
                    )
                    continue

                pend_normales.append({
                    "agente": agente, "agente_desc": nombre_raw, "tipo": tipo,
                    "fecha": date(cfg["anio"], 1, 1), "cantidad": round(valor),
                    "hoja": cfg["sheet"], "anio_hoja": cfg["anio"], "fila": fila_excel, "col": col,
                })

    def _guardar_licencia(self, item, avisos, stats):
        agente, tipo, fecha, cantidad = item["agente"], item["tipo"], item["fecha"], item["cantidad"]
        motivo = (
            f"{TAG} Total anual registrado en la hoja '{item['hoja']}' para {item['anio_hoja']}. "
            "Sin fecha exacta de goce; ver la planilla original para el detalle dia a dia."
        )

        existente = LicenciaPermiso.objects.filter(
            licenciapermiso_agente=agente, licenciapermiso_tipo=tipo,
            licenciapermiso_fecha_desde=fecha, licenciapermiso_motivo__startswith=TAG,
        ).first()

        if existente:
            cambio = existente.licenciapermiso_cantidad != cantidad
            existente.licenciapermiso_cantidad = cantidad
            existente.licenciapermiso_motivo = motivo
            licencia = existente
        else:
            cambio = True
            licencia = LicenciaPermiso(
                licenciapermiso_agente=agente, licenciapermiso_tipo=tipo,
                licenciapermiso_fecha_desde=fecha, licenciapermiso_fecha_otorgamiento=fecha,
                licenciapermiso_cantidad=cantidad, licenciapermiso_motivo=motivo,
            )

        if not cambio:
            stats["sin_cambios"] += 1
            return

        try:
            licencia.full_clean()
        except ValidationError as exc:
            stats["rechazados"] += 1
            _registrar_aviso(
                avisos, item["hoja"], item["fila"], item["agente_desc"], item["col"],
                f"'{item['agente_desc']}': no se pudo guardar {tipo.tipolicenciapermiso_nombre} "
                f"{item['anio_hoja']} (cantidad={cantidad}): {exc}",
            )
            return

        # Se guarda siempre, incluso en dry-run: la validacion del Art. 10
        # (segunda pasada) depende de que las "Anual" de la primera pasada
        # esten realmente persistidas (aunque sea solo dentro de esta misma
        # transaccion) para calcular bien el cupo ya comprometido. El
        # rollback de dry-run (ver handle()) deshace todo esto al final.
        licencia.save()
        stats["actualizados" if existente else "creados"] += 1

    def _exportar_errores(self, headers_por_hoja, filas_vistas, avisos, ruta):
        """Vuelca, en una hoja por cada hoja de origen, las filas que
        generaron algun aviso, resaltando las celdas puntuales que lo
        causaron, para revision manual."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        errores_por_hoja = defaultdict(lambda: defaultdict(lambda: {"mensajes": [], "cols": set()}))
        for aviso in avisos:
            entry = errores_por_hoja[aviso["hoja"]][aviso["fila"]]
            entry["mensajes"].append(aviso["msg"])
            entry["cols"].update(aviso["cols"])

        resaltado = PatternFill(fill_type="solid", start_color="FFF4B6C2", end_color="FFF4B6C2")
        negrita = Font(bold=True)

        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)

        nombres_usados = set()
        for hoja, filas in errores_por_hoja.items():
            nombre_hoja = hoja.strip()[:31] or "Errores"
            base, i = nombre_hoja, 1
            while nombre_hoja in nombres_usados:
                sufijo = f" ({i})"
                nombre_hoja = base[:31 - len(sufijo)] + sufijo
                i += 1
            nombres_usados.add(nombre_hoja)

            hoja_out = wb_out.create_sheet(nombre_hoja)
            header = headers_por_hoja.get(hoja, [])
            # El xlsx de origen trae columnas de formato basura sin datos
            # mas alla del header real (ej. hasta la columna 160 en alguna
            # hoja): se recorta a lo que realmente tiene contenido, tomando
            # el mayor entre el largo del header (sin el relleno de None al
            # final) y la ultima columna con dato en alguna fila con aviso.
            largo_header = len(header)
            while largo_header and _blank(header[largo_header - 1]):
                largo_header -= 1
            ultima_col_dato = -1
            for fila_excel in filas:
                row = filas_vistas.get((hoja, fila_excel), ())
                for i, v in enumerate(row):
                    if not _blank(v):
                        ultima_col_dato = max(ultima_col_dato, i)
            ncols = max(largo_header, ultima_col_dato + 1)
            encabezado = [(header[i] if i < len(header) else None) or f"Col {i}" for i in range(ncols)] + ["Fila xlsx", "Avisos"]
            hoja_out.append(encabezado)
            for celda in hoja_out[1]:
                celda.font = negrita

            for fila_excel in sorted(filas):
                info = filas[fila_excel]
                row = filas_vistas.get((hoja, fila_excel), ())
                valores = list(row[:ncols]) + [None] * (ncols - len(row))
                valores += [fila_excel, "; ".join(info["mensajes"])]
                hoja_out.append(valores)

                fila_hoja = hoja_out.max_row
                for col in info["cols"]:
                    hoja_out.cell(row=fila_hoja, column=col + 1).fill = resaltado
                hoja_out.cell(row=fila_hoja, column=ncols + 2).fill = resaltado

        wb_out.save(ruta)
        return ruta
