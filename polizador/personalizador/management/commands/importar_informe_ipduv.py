import re
import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.validators import get_cuil
from carga.models import Localidad
from personalizador.models import (
    Agente,
    ActividadEspecifica,
    ApartadoCargo,
    CEIC,
    Departamento,
    DenominacionCargo,
    Direccion,
    Directorio,
    Gerencia,
    GeneroAgente,
    GrupoCargo,
    Oficina,
    TituloProfesional,
)

logger = logging.getLogger(__name__)

DEFAULT_PATH = settings.BASE_DIR.parent / "env" / ".snipets" / "personal_gral.xlsx"

SHEET_NAME = "PERSONAL"

# Columnas de la hoja "PERSONAL" de personal_gral.xlsx (0-indexadas). A
# diferencia del viejo informe_ipduv.xlsx, esta hoja NO viene con el header
# renombrado a mano: usa los rotulos originales en español, repartidos en 2
# filas (2 y 3) con celdas combinadas (ej. "APORTES DE LEY 6039" en la fila 2
# combinada sobre "AÑO"/"MES"/"DIA"/"RESOLUCION" en la fila 3), y hay rotulos
# duplicados (dos columnas "C.U.OF.", "MES"/"AÑO"/"DIA" repetidos en distintos
# bloques) asi que no se puede armar un dict desde el header sin perder
# columnas: se hardcodean las posiciones, verificadas contra el archivo real.
#
# Ademas, el rotulo de la columna 17 es enganoso: dice "CATEGORIA" pero sus
# valores ("PROFESIONAL  5", "ADMINISTRATIVO 6", ...) son en realidad los
# mismos que "denominacion_cargo" en el informe viejo. El archivo no trae
# ninguna columna con el formato viejo de categoria ("3 - Personal
# Administrativo y Tecnico"), asi que el campo Agente.categoria ya no se
# puede completar desde esta fuente y este comando no lo toca.
#
# Las columnas 8 ("resolucion contrato de servicio tipo"), 14
# ("resolucion_titulo_tipo") y 40 ("Tipo", dentro del bloque "APORTES DE LEY
# 6039") se agregaron despues, a mano, para completar
# Agente.instrumento_contrato_de_servicio_tipo,
# Agente.instrumento_bonificacion_tipo y Agente.aportes_ley_resolucion_tipo
# respectivamente: vienen ya codificadas con los mismos valores que los
# choices del modelo (RES/DEC, RES/DEC y DEC/RES/DIS), sin encabezado en la
# fila 2.
#
# Los bloques "ANTIGÜEDAD BONIFICABLE" (AÑOS/MES/DIAS/FECHA DE SISTEMA) y
# "AÑOS TOTALES" son totales ya calculados (equivalentes a la property
# Agente.antiguedad), no datos de carga: no se importan para no pisar ese
# calculo con un valor congelado a la fecha del reporte.
#
# La columna "CARGO INTERNO" (designacion transitoria) contiene texto libre
# de cargos/roles (ej. "GERENTE RECURSOS FINANCIEROS", "COORDINADOR DE
# CEREMONIAL Y PROTOCOLO") que no matchea contra los nombres de las unidades
# organizativas (Departamento/Direccion/Gerencia/Directorio): en una prueba
# contra la base no matcheo ninguna de las 57 filas con valor, asi que no se
# importa (Agente.cargo_interno / instrumento_cargo_interno quedan afuera).
COL_N_LEGAJO = 0
COL_SEXO = 1
COL_DNI = 3
COL_FECHA_NACIMIENTO = 5
COL_FECHA_INGRESO = 7
COL_INSTRUMENTO_CONTRATO_SERVICIO_TIPO = 8
COL_N_RESOLUCION_CONTRATO_SERVICIO = 9
COL_FECHA_PASE_A_PLANTA = 10
COL_N_DECRETO = 11
COL_TITULO_GRADO = 12
COL_TITULO_NOMBRE = 13
COL_INSTRUMENTO_BONIFICACION_TIPO = 14
COL_N_RESOLUCION_BONIFICACION = 15
COL_PORCENTAJE_BONIFICACION = 16
COL_DENOMINACION_CARGO = 17
COL_APARTADO = 18
COL_CEIC = 19
COL_GRUPO = 20
COL_ACTIVIDAD_CENTRAL = 21
COL_ACTIVIDAD_ESPECIFICA = 22
COL_CUOF_DEPARTAMENTO = 25
COL_DEPARTAMENTO_NOMBRE = 26
COL_N_DECRETO_TRANSFERENCIA = 27
COL_DOMICILIO_DIRECCION = 30
COL_DOMICILIO_BARRIO = 31
COL_DOMICILIO_LOCALIDAD = 32
COL_APORTES_LEY_ANIOS = 37
COL_APORTES_LEY_MESES = 38
COL_APORTES_LEY_DIAS = 39
COL_APORTES_LEY_TIPO = 40
COL_APORTES_LEY_RESOLUCION = 41
COL_APORTES_ANSES_ANIOS = 42
COL_APORTES_ANSES_MESES = 43
COL_APORTES_ANSES_DIAS = 44

FIRST_DATA_ROW = 4

SEXO_A_GENERO = {"M": "Masculino", "F": "Femenino"}

# Estas columnas ("resolucion contrato de servicio tipo", "resolucion_titulo_tipo"
# y el "Tipo" dentro del bloque "APORTES DE LEY 6039") se agregaron a mano al
# xlsx despues de los campos *_tipo del modelo, y ya vienen codificadas con
# los mismos valores de los choices (RES/DEC, DEC/RES/DIS) - no necesitan
# mapeo, solo validacion.
INSTRUMENTO_CONTRATO_SERVICIO_TIPO_CODIGOS = {"RES", "DEC"}
INSTRUMENTO_BONIFICACION_TIPO_CODIGOS = {"RES", "DEC"}
APORTES_LEY_RESOLUCION_TIPO_CODIGOS = {"DEC", "RES", "DIS"}

# Fallback si "APORTES DE LEY 6039 - Tipo" viene vacio: se infiere del texto
# de la columna RESOLUCION (prefijos Decreto/Resolucion/Disposicion).
TIPO_INSTRUMENTO_PATRONES = (
    (re.compile(r"\bdec|\bdto\b", re.I), "DEC"),
    (re.compile(r"\bres", re.I), "RES"),
    (re.compile(r"\bdis", re.I), "DIS"),
)

# Rotulos legibles para las columnas que este comando efectivamente lee, para
# el encabezado del xlsx de errores (ver _exportar_errores). Las columnas sin
# entrada se listan como "Col <n>".
COL_LABELS = {
    COL_N_LEGAJO: "Nº",
    COL_SEXO: "SEXO",
    COL_DNI: "D.N.I.",
    COL_FECHA_NACIMIENTO: "F.NAC.",
    COL_FECHA_INGRESO: "FECH.ING.",
    COL_INSTRUMENTO_CONTRATO_SERVICIO_TIPO: "resolucion contrato de servicio tipo",
    COL_N_RESOLUCION_CONTRATO_SERVICIO: "Nº RESOLUCION DE CONTRATO DE SERVICIO",
    COL_FECHA_PASE_A_PLANTA: "F. PASE A PLANTA",
    COL_N_DECRETO: "Nº DE DECRETO",
    COL_TITULO_GRADO: "TITULO",
    COL_TITULO_NOMBRE: "TIPO",
    COL_INSTRUMENTO_BONIFICACION_TIPO: "resolucion_titulo_tipo",
    COL_N_RESOLUCION_BONIFICACION: "Nº RESOLUCION TITULO",
    COL_PORCENTAJE_BONIFICACION: "PORCENTAJE",
    COL_DENOMINACION_CARGO: "CATEGORIA (denominacion_cargo)",
    COL_APARTADO: "APARTADO",
    COL_CEIC: "C.E.I.C",
    COL_GRUPO: "GRUPO",
    COL_ACTIVIDAD_CENTRAL: "ACT. CENTRAL",
    COL_ACTIVIDAD_ESPECIFICA: "ACT. ESPECIFICA",
    COL_CUOF_DEPARTAMENTO: "C.U.OF. (departamento)",
    COL_DEPARTAMENTO_NOMBRE: "DEPARTAMENTO Y/O DIRECCIONES",
    COL_N_DECRETO_TRANSFERENCIA: "TRASFERENCIAS Nº DECRETO",
    COL_DOMICILIO_DIRECCION: "DIRECCION",
    COL_DOMICILIO_BARRIO: "BARRIO-VILLA",
    COL_DOMICILIO_LOCALIDAD: "LOCALIDAD",
    COL_APORTES_LEY_ANIOS: "APORTES LEY 6039 - AÑOS",
    COL_APORTES_LEY_MESES: "APORTES LEY 6039 - MES",
    COL_APORTES_LEY_DIAS: "APORTES LEY 6039 - DIA",
    COL_APORTES_LEY_TIPO: "APORTES LEY 6039 - Tipo",
    COL_APORTES_LEY_RESOLUCION: "APORTES LEY 6039 - RESOLUCION",
    COL_APORTES_ANSES_ANIOS: "ANSES - AÑO",
    COL_APORTES_ANSES_MESES: "ANSES - MES",
    COL_APORTES_ANSES_DIAS: "ANSES - DIA",
}


def _blank(raw):
    return raw is None or (isinstance(raw, str) and raw.strip() == "")


def _parse_codigo_nombre(raw):
    """'02 - ADMINISTRACION Y RECURSOS HUMANOS' -> (2, 'ADMINISTRACION Y RECURSOS HUMANOS')"""
    texto = str(raw).strip()
    codigo_str, _, nombre_str = texto.partition("-")
    codigo_str = codigo_str.strip()
    if not codigo_str.isdigit():
        return None, None
    return int(codigo_str), nombre_str.strip()


def _normalize_text(raw):
    return " ".join(str(raw).split())


def _parse_dni(raw):
    if _blank(raw):
        return None
    digitos = re.sub(r"\D", "", str(raw))
    return int(digitos) if digitos else None


def _parse_leading_int(raw):
    if _blank(raw):
        return None
    match = re.match(r"\d+", str(raw).strip())
    return match.group() if match else None


def _parse_fecha(raw):
    return raw.date() if hasattr(raw, "date") else raw


def _parse_positive_int(raw):
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _fit_charfield(raw, max_length):
    valor = _normalize_text(raw)
    return valor if len(valor) <= max_length else None


def _parse_tipo_instrumento(raw):
    texto = str(raw)
    for patron, codigo in TIPO_INSTRUMENTO_PATRONES:
        if patron.search(texto):
            return codigo
    return None


def _parse_choice(raw, codigos_validos):
    valor = _normalize_text(raw).upper()
    return valor if valor in codigos_validos else None


def _registrar_aviso(avisos, fila, dni, cols, msg, silencioso=False):
    """Guarda un aviso con la fila/columnas del xlsx a las que se refiere,
    para poder resaltarlas despues en el xlsx de errores (ver
    _exportar_errores). Si silencioso=True, el aviso va igual al xlsx pero
    no se imprime por consola (ver el filtro en handle())."""
    if cols is None:
        cols = []
    elif isinstance(cols, int):
        cols = [cols]
    avisos.append({"fila": fila, "dni": dni, "cols": list(cols), "msg": msg, "silencioso": silencioso})


class Command(BaseCommand):
    help = (
        "Importa la hoja 'PERSONAL' de personal_gral.xlsx: matchea Agentes "
        "existentes por DNI y completa sexo, fecha_nacimiento, "
        "fecha_ingreso, instrumento_contrato_de_servicio + "
        "instrumento_contrato_de_servicio_tipo (leido de la columna "
        "'resolucion contrato de servicio tipo'), fecha_pase_a_planta, n_decreto, "
        "instrumento_bonificacion, instrumento_bonificacion_tipo (leido de la "
        "columna 'resolucion_titulo_tipo'), porcentaje_bonificacion, "
        "titulo_profesional, denominacion_cargo, apartado, ceic, grupo, "
        "activdad_central, actividad_especifica, oficina (resuelta a partir "
        "del C.U.O.F. de departamento/direccion contra el organigrama "
        "cargado por fill_data.py + crear_oficinas.py), "
        "n_decreto_transferencia_definitiva, domicilio_direccion, "
        "domicilio_barrio, domicilio_localidad y los aportes de ley "
        "6039/ANSES (años/meses/dias + resolucion y tipo de instrumento "
        "-Decreto/Resolucion/Disposicion-, leido de la columna 'Tipo' del "
        "bloque de aportes de ley 6039 o, si viene vacia, inferido del texto "
        "de la resolucion). "
        "El CUIL no se lee de la columna del xlsx (viene con formato "
        "inconsistente): se calcula con core.validators.get_cuil() a partir "
        "del DNI y el SEXO de cada fila. "
        "Ademas marca Agente.activo=True para los Agentes matcheados por DNI "
        "presente en el xlsx, y Agente.activo=False para todos los demas "
        "Agentes de la base que no aparezcan en el listado (con DNI valido). "
        "Tambien marca Agente.con_errores=True si la fila del agente genero "
        "algun aviso (incluidos los de TituloProfesional silenciados de la "
        "consola) y con_errores=False si no genero ninguno, para poder "
        "filtrar despues los agentes con carga incompleta desde el listado. "
        "No crea Agentes nuevos ni filas de catalogo faltantes (Categoria, "
        "DenominacionCargo, ApartadoCargo, CEIC, GrupoCargo, "
        "ActividadEspecifica, Oficina, GeneroAgente, TituloProfesional, "
        "Localidad): si un valor del xlsx no tiene correspondencia en la "
        "base, se salta ese campo puntual para ese agente y se informa por "
        "consola para revision manual (excepto los de TituloProfesional no "
        "encontrado, que salen muy seguido por variantes de genero/typos y se "
        "silencian de la consola -solo quedan en el xlsx de errores-). Los "
        "campos vacios en el xlsx tambien se saltean (no se pisan datos "
        "existentes por una celda en blanco). Si hubo avisos, ademas se "
        "exporta un xlsx (--errores-file) con las filas afectadas y las "
        "celdas con error resaltadas (incluye tambien los avisos "
        "silenciados de la consola)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", default=str(DEFAULT_PATH),
            help="Ruta al xlsx a importar (default: %(default)s)",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="No guarda cambios en la base de datos, solo informa.",
        )
        parser.add_argument(
            "--errores-file", default=None,
            help="Ruta donde guardar el xlsx con las filas que tuvieron avisos "
                 "(default: el archivo de entrada con sufijo '_errores').",
        )

    def handle(self, *args, **options):
        import openpyxl

        path = options["file"]
        dry_run = options["dry_run"]
        verbosity = options.get("verbosity", 1)

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontro el archivo: {path}")

        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"La hoja '{SHEET_NAME}' no existe en {path}.")
        ws = wb[SHEET_NAME]

        stats = {"actualizados": 0, "sin_cambios": 0, "no_encontrados": 0, "dni_invalido": 0, "marcados_inactivos": 0}
        avisos = []
        filas_vistas = {}
        dnis_en_listado = set()

        with transaction.atomic():
            for fila_excel, row in enumerate(
                ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW,
            ):
                if row[COL_N_LEGAJO] is None and row[COL_DNI] is None:
                    continue

                filas_vistas[fila_excel] = row

                dni = _parse_dni(row[COL_DNI])
                if dni is None:
                    stats["dni_invalido"] += 1
                    _registrar_aviso(
                        avisos, fila_excel, None, COL_DNI,
                        f"Fila con N={row[COL_N_LEGAJO]!r}: D.N.I. invalido ({row[COL_DNI]!r}), fila salteada.",
                    )
                    continue

                dnis_en_listado.add(dni)

                try:
                    agente = Agente.objects.get(dni=dni)
                except Agente.DoesNotExist:
                    stats["no_encontrados"] += 1
                    _registrar_aviso(
                        avisos, fila_excel, dni, COL_DNI,
                        f"DNI {dni}: no existe ningun Agente con ese DNI, fila salteada.",
                    )
                    continue

                avisos_antes = len(avisos)

                updates = self._resolver_campos(row, dni, fila_excel, avisos)
                updates["activo"] = True

                titulo = self._resolver_titulo(row, dni, fila_excel, avisos)
                titulo_nuevo = titulo is not None and not agente.titulo_profesional.filter(pk=titulo.pk).exists()

                updates["con_errores"] = len(avisos) > avisos_antes

                changed_fields = []
                for field, value in updates.items():
                    if getattr(agente, field) != value:
                        setattr(agente, field, value)
                        changed_fields.append(field)

                if titulo_nuevo:
                    changed_fields.append("titulo_profesional")

                if changed_fields:
                    stats["actualizados"] += 1
                    if not dry_run:
                        agente.save()
                        if titulo_nuevo:
                            agente.titulo_profesional.add(titulo)
                    if verbosity >= 2:
                        self.stdout.write(f"DNI {dni}: actualizados {', '.join(changed_fields)}")
                else:
                    stats["sin_cambios"] += 1

            stats["marcados_inactivos"] = Agente.objects.exclude(dni__in=dnis_en_listado).filter(activo=True).update(activo=False)

            if dry_run:
                transaction.set_rollback(True)

        if verbosity >= 1:
            self.stdout.write("")
            if dry_run:
                self.stdout.write(self.style.NOTICE("Modo dry-run: no se guardo ningun cambio en la base de datos."))
            self.stdout.write(self.style.MIGRATE_HEADING("Resumen:"))
            self.stdout.write(self.style.SUCCESS(f"    Agentes actualizados: {stats['actualizados']}"))
            self.stdout.write(f"    Agentes sin cambios: {stats['sin_cambios']}")
            self.stdout.write(self.style.WARNING(f"    DNI no encontrado en la base: {stats['no_encontrados']}"))
            self.stdout.write(self.style.WARNING(f"    D.N.I. invalido en el xlsx: {stats['dni_invalido']}"))
            self.stdout.write(self.style.WARNING(f"    Agentes marcados inactivos (no estan en el listado): {stats['marcados_inactivos']}"))

        avisos_a_mostrar = [a for a in avisos if not a["silencioso"]]
        avisos_silenciados = len(avisos) - len(avisos_a_mostrar)
        if avisos and verbosity >= 1:
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(
                f"Avisos ({len(avisos_a_mostrar)}"
                + (f", +{avisos_silenciados} de TituloProfesional silenciados (ver el xlsx de errores)" if avisos_silenciados else "")
                + "):"
            ))
            for aviso in avisos_a_mostrar:
                self.stdout.write(f"    {self.style.WARNING(aviso['msg'])}")

        if avisos:
            ruta_errores = options["errores_file"]
            if not ruta_errores:
                origen = Path(path)
                ruta_errores = str(origen.with_name(f"{origen.stem}_errores.xlsx"))
            self._exportar_errores(filas_vistas, avisos, ruta_errores)
            if verbosity >= 1:
                self.stdout.write(self.style.NOTICE(f"Filas con avisos exportadas a: {ruta_errores}"))

    def _exportar_errores(self, filas_vistas, avisos, ruta):
        """Vuelca las filas del xlsx que generaron algun aviso, resaltando
        las celdas puntuales que lo causaron, para revision manual."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        errores_por_fila = {}
        for aviso in avisos:
            entry = errores_por_fila.setdefault(aviso["fila"], {"mensajes": [], "cols": set()})
            entry["mensajes"].append(aviso["msg"])
            entry["cols"].update(aviso["cols"])

        ncols = max((len(row) for row in filas_vistas.values()), default=0)

        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "Errores"

        resaltado = PatternFill(fill_type="solid", start_color="FFF4B6C2", end_color="FFF4B6C2")
        negrita = Font(bold=True)

        encabezado = [COL_LABELS.get(i, f"Col {i}") for i in range(ncols)] + ["Fila xlsx", "Avisos"]
        hoja.append(encabezado)
        for celda in hoja[1]:
            celda.font = negrita

        for fila_excel in sorted(errores_por_fila):
            info = errores_por_fila[fila_excel]
            row = filas_vistas.get(fila_excel, ())
            valores = list(row) + [None] * (ncols - len(row))
            valores += [fila_excel, "; ".join(info["mensajes"])]
            hoja.append(valores)

            fila_hoja = hoja.max_row
            for col in info["cols"]:
                hoja.cell(row=fila_hoja, column=col + 1).fill = resaltado
            hoja.cell(row=fila_hoja, column=ncols + 2).fill = resaltado

        wb.save(ruta)
        return ruta

    def _resolver_campos(self, row, dni, fila_excel, avisos):
        updates = {}

        if not _blank(row[COL_N_LEGAJO]):
            try:
                updates["n_legajo"] = int(row[COL_N_LEGAJO])
            except (TypeError, ValueError):
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_N_LEGAJO,
                    f"DNI {dni}: N de legajo invalido ({row[COL_N_LEGAJO]!r}), se salteo el campo.",
                )

        if _blank(row[COL_SEXO]):
            _registrar_aviso(
                avisos, fila_excel, dni, COL_SEXO,
                f"DNI {dni}: SEXO vacio en el xlsx, no se pudo calcular CUIL.",
            )
        else:
            valor = str(row[COL_SEXO]).strip().upper()
            nombre_genero = SEXO_A_GENERO.get(valor)
            genero = GeneroAgente.objects.filter(generoagente_nombre__iexact=nombre_genero).first() if nombre_genero else None
            if genero:
                updates["sexo"] = genero
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_SEXO,
                    f"DNI {dni}: SEXO '{row[COL_SEXO]!r}' no reconocido o no existe GeneroAgente correspondiente, se salteo el campo.",
                )

            if nombre_genero is None:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_SEXO,
                    f"DNI {dni}: no se pudo calcular CUIL porque SEXO '{row[COL_SEXO]!r}' no es M/F.",
                )
            else:
                try:
                    cuil_crudo = get_cuil(str(dni), valor)
                    updates["cuil"] = f"{cuil_crudo[:2]}-{cuil_crudo[2:10]}-{cuil_crudo[10:]}"
                except ValidationError as exc:
                    _registrar_aviso(
                        avisos, fila_excel, dni, COL_DNI,
                        f"DNI {dni}: no se pudo calcular CUIL ({exc}), se salteo el campo.",
                    )

        if not _blank(row[COL_FECHA_NACIMIENTO]):
            updates["fecha_nacimiento"] = _parse_fecha(row[COL_FECHA_NACIMIENTO])

        if not _blank(row[COL_FECHA_INGRESO]):
            updates["fecha_ingreso"] = _parse_fecha(row[COL_FECHA_INGRESO])

        if not _blank(row[COL_N_RESOLUCION_CONTRATO_SERVICIO]):
            valor = _fit_charfield(row[COL_N_RESOLUCION_CONTRATO_SERVICIO], 13)
            if valor:
                updates["instrumento_contrato_de_servicio"] = valor
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_N_RESOLUCION_CONTRATO_SERVICIO,
                    f"DNI {dni}: Nº de resolucion de contrato de servicio demasiado largo ({row[COL_N_RESOLUCION_CONTRATO_SERVICIO]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_INSTRUMENTO_CONTRATO_SERVICIO_TIPO]):
            tipo = _parse_choice(row[COL_INSTRUMENTO_CONTRATO_SERVICIO_TIPO], INSTRUMENTO_CONTRATO_SERVICIO_TIPO_CODIGOS)
            if tipo:
                updates["instrumento_contrato_de_servicio_tipo"] = tipo
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_INSTRUMENTO_CONTRATO_SERVICIO_TIPO,
                    f"DNI {dni}: resolucion contrato de servicio tipo invalido ({row[COL_INSTRUMENTO_CONTRATO_SERVICIO_TIPO]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_FECHA_PASE_A_PLANTA]):
            updates["fecha_pase_a_planta"] = _parse_fecha(row[COL_FECHA_PASE_A_PLANTA])

        if not _blank(row[COL_N_DECRETO]):
            valor = _fit_charfield(row[COL_N_DECRETO], 10)
            if valor:
                updates["n_decreto"] = valor
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_N_DECRETO,
                    f"DNI {dni}: Nº de decreto demasiado largo ({row[COL_N_DECRETO]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_N_RESOLUCION_BONIFICACION]):
            valor = _fit_charfield(row[COL_N_RESOLUCION_BONIFICACION], 13)
            if valor:
                updates["instrumento_bonificacion"] = valor
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_N_RESOLUCION_BONIFICACION,
                    f"DNI {dni}: Nº de resolucion de titulo demasiado largo ({row[COL_N_RESOLUCION_BONIFICACION]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_INSTRUMENTO_BONIFICACION_TIPO]):
            tipo = _parse_choice(row[COL_INSTRUMENTO_BONIFICACION_TIPO], INSTRUMENTO_BONIFICACION_TIPO_CODIGOS)
            if tipo:
                updates["instrumento_bonificacion_tipo"] = tipo
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_INSTRUMENTO_BONIFICACION_TIPO,
                    f"DNI {dni}: resolucion_titulo_tipo invalido ({row[COL_INSTRUMENTO_BONIFICACION_TIPO]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_PORCENTAJE_BONIFICACION]):
            try:
                updates["porcentaje_bonificacion"] = Decimal(str(row[COL_PORCENTAJE_BONIFICACION]))
            except InvalidOperation:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_PORCENTAJE_BONIFICACION,
                    f"DNI {dni}: PORCENTAJE invalido ({row[COL_PORCENTAJE_BONIFICACION]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_DENOMINACION_CARGO]):
            normalizado = _normalize_text(row[COL_DENOMINACION_CARGO])
            denominacion = DenominacionCargo.objects.filter(denominacion__iexact=normalizado).first()
            if denominacion:
                updates["denominacion_cargo"] = denominacion
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_DENOMINACION_CARGO,
                    f"DNI {dni}: no existe DenominacionCargo '{normalizado}', se salteo el campo.",
                )

        if not _blank(row[COL_APARTADO]):
            valor = str(row[COL_APARTADO]).strip().upper()
            apartado = ApartadoCargo.objects.filter(apartadocargo_denominacion=valor).first()
            if apartado:
                updates["apartado"] = apartado
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_APARTADO,
                    f"DNI {dni}: no existe ApartadoCargo '{valor}', se salteo el campo.",
                )

        if not _blank(row[COL_CEIC]):
            valor = _parse_leading_int(row[COL_CEIC])
            ceic = CEIC.objects.filter(ceic=valor).first() if valor else None
            if ceic:
                updates["ceic"] = ceic
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_CEIC,
                    f"DNI {dni}: no existe CEIC '{row[COL_CEIC]!r}', se salteo el campo.",
                )

        if not _blank(row[COL_GRUPO]):
            try:
                valor = int(row[COL_GRUPO])
            except (TypeError, ValueError):
                valor = None
            grupo = GrupoCargo.objects.filter(grupo_numero=valor).first() if valor is not None else None
            if grupo:
                updates["grupo"] = grupo
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_GRUPO,
                    f"DNI {dni}: no existe GrupoCargo '{row[COL_GRUPO]!r}', se salteo el campo.",
                )

        if not _blank(row[COL_ACTIVIDAD_CENTRAL]):
            try:
                updates["activdad_central"] = str(int(row[COL_ACTIVIDAD_CENTRAL]))
            except (TypeError, ValueError):
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_ACTIVIDAD_CENTRAL,
                    f"DNI {dni}: ACT. CENTRAL invalida ({row[COL_ACTIVIDAD_CENTRAL]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_ACTIVIDAD_ESPECIFICA]):
            codigo, _nombre = _parse_codigo_nombre(row[COL_ACTIVIDAD_ESPECIFICA])
            actividad = ActividadEspecifica.objects.filter(actividad_especifica_codigo=codigo).first() if codigo is not None else None
            if actividad:
                updates["actividad_especifica"] = actividad
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_ACTIVIDAD_ESPECIFICA,
                    f"DNI {dni}: no existe ActividadEspecifica con codigo {codigo!r} (xlsx: {row[COL_ACTIVIDAD_ESPECIFICA]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_CUOF_DEPARTAMENTO]):
            oficina = self._resolver_oficina(row, dni, fila_excel, avisos)
            if oficina is not None:
                updates["oficina"] = oficina

        if not _blank(row[COL_N_DECRETO_TRANSFERENCIA]):
            valor = _fit_charfield(row[COL_N_DECRETO_TRANSFERENCIA], 10)
            if valor:
                updates["n_decreto_transferencia_definitiva"] = valor
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_N_DECRETO_TRANSFERENCIA,
                    f"DNI {dni}: Nº de decreto de transferencia demasiado largo ({row[COL_N_DECRETO_TRANSFERENCIA]!r}), se salteo el campo.",
                )

        if not _blank(row[COL_DOMICILIO_DIRECCION]):
            updates["domicilio_direccion"] = _normalize_text(row[COL_DOMICILIO_DIRECCION])

        if not _blank(row[COL_DOMICILIO_BARRIO]):
            updates["domicilio_barrio"] = _normalize_text(row[COL_DOMICILIO_BARRIO])

        if not _blank(row[COL_DOMICILIO_LOCALIDAD]):
            nombre_localidad = _normalize_text(row[COL_DOMICILIO_LOCALIDAD]).strip(". ")
            localidad = Localidad.objects.filter(localidad_nombre__iexact=nombre_localidad).first()
            if localidad:
                updates["domicilio_localidad"] = localidad
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_DOMICILIO_LOCALIDAD,
                    f"DNI {dni}: no existe Localidad '{nombre_localidad}', se salteo el campo.",
                )

        for col, field, label in (
            (COL_APORTES_LEY_ANIOS, "aportes_ley_resolucion_anios", "años de aportes de ley 6039"),
            (COL_APORTES_LEY_MESES, "aportes_ley_resolucion_meses", "meses de aportes de ley 6039"),
            (COL_APORTES_LEY_DIAS, "aportes_ley_resolucion_dias", "dias de aportes de ley 6039"),
            (COL_APORTES_ANSES_ANIOS, "aportes_anses_anios", "años de aportes ANSES"),
            (COL_APORTES_ANSES_MESES, "aportes_anses_meses", "meses de aportes ANSES"),
            (COL_APORTES_ANSES_DIAS, "aportes_anses_dias", "dias de aportes ANSES"),
        ):
            if not _blank(row[col]):
                valor = _parse_positive_int(row[col])
                if valor is not None:
                    updates[field] = valor
                else:
                    _registrar_aviso(
                        avisos, fila_excel, dni, col,
                        f"DNI {dni}: valor invalido para {label} ({row[col]!r}), se salteo el campo.",
                    )

        if not _blank(row[COL_APORTES_LEY_RESOLUCION]):
            valor = _fit_charfield(row[COL_APORTES_LEY_RESOLUCION], 13)
            if valor:
                updates["aportes_ley_resolucion"] = valor
                if not _blank(row[COL_APORTES_LEY_TIPO]):
                    tipo = _parse_choice(row[COL_APORTES_LEY_TIPO], APORTES_LEY_RESOLUCION_TIPO_CODIGOS)
                    if tipo:
                        updates["aportes_ley_resolucion_tipo"] = tipo
                    else:
                        _registrar_aviso(
                            avisos, fila_excel, dni, COL_APORTES_LEY_TIPO,
                            f"DNI {dni}: APORTES LEY 6039 - Tipo invalido ({row[COL_APORTES_LEY_TIPO]!r}), se salteo el campo.",
                        )
                else:
                    tipo = _parse_tipo_instrumento(row[COL_APORTES_LEY_RESOLUCION])
                    if tipo:
                        updates["aportes_ley_resolucion_tipo"] = tipo
                    else:
                        _registrar_aviso(
                            avisos, fila_excel, dni, COL_APORTES_LEY_RESOLUCION,
                            f"DNI {dni}: no se pudo determinar el tipo de instrumento (Decreto/Resolucion/"
                            f"Disposicion) de aportes de ley 6039 a partir de '{row[COL_APORTES_LEY_RESOLUCION]!r}', "
                            f"se salteo aportes_ley_resolucion_tipo.",
                        )
            else:
                _registrar_aviso(
                    avisos, fila_excel, dni, COL_APORTES_LEY_RESOLUCION,
                    f"DNI {dni}: resolucion de aportes de ley 6039 demasiado larga ({row[COL_APORTES_LEY_RESOLUCION]!r}), se salteo el campo.",
                )

        return updates

    def _resolver_titulo(self, row, dni, fila_excel, avisos):
        grado_raw, nombre_raw = row[COL_TITULO_GRADO], row[COL_TITULO_NOMBRE]
        if _blank(grado_raw) or _blank(nombre_raw):
            return None

        grado = _normalize_text(grado_raw)
        nombre = _normalize_text(nombre_raw)
        titulo = TituloProfesional.objects.filter(
            tituloprofesional_grado__iexact=grado, tituloprofesional_nombre__iexact=nombre,
        ).first()
        if not titulo:
            _registrar_aviso(
                avisos, fila_excel, dni, [COL_TITULO_GRADO, COL_TITULO_NOMBRE],
                f"DNI {dni}: no existe TituloProfesional '{grado} - {nombre}', se salteo.",
                silencioso=True,
            )
        return titulo

    def _resolver_oficina(self, row, dni, fila_excel, avisos):
        try:
            numero = int(row[COL_CUOF_DEPARTAMENTO])
        except (TypeError, ValueError):
            _registrar_aviso(
                avisos, fila_excel, dni, COL_CUOF_DEPARTAMENTO,
                f"DNI {dni}: C.U.O.F. de departamento/direccion invalido ({row[COL_CUOF_DEPARTAMENTO]!r}), se salteo oficina.",
            )
            return None
        cuof = f"10-{numero}-0"
        nombre_xlsx = row[COL_DEPARTAMENTO_NOMBRE]

        departamento = Departamento.objects.filter(departamento_cuof=cuof).first()
        if departamento:
            oficina = Oficina.objects.filter(cargo_departamento=departamento).first()
            nivel, unidad = "Departamento", departamento.departamento_nombre
        else:
            direccion = Direccion.objects.filter(direccion_cuof=cuof).first()
            if direccion:
                oficina = Oficina.objects.filter(cargo_direccion=direccion, cargo_departamento__isnull=True).first()
                nivel, unidad = "Direccion", direccion.direccion_nombre
            else:
                gerencia = Gerencia.objects.filter(gerencia_cuof=cuof).first()
                if gerencia:
                    oficina = Oficina.objects.filter(
                        cargo_gerencia=gerencia, cargo_direccion__isnull=True, cargo_departamento__isnull=True,
                    ).first()
                    nivel, unidad = "Gerencia", gerencia.gerencia_nombre
                else:
                    directorio = Directorio.objects.filter(directorio_cuof=cuof).first()
                    if directorio:
                        oficina = Oficina.objects.filter(
                            cargo_directorio=directorio, cargo_gerencia__isnull=True,
                            cargo_direccion__isnull=True, cargo_departamento__isnull=True,
                        ).first()
                        nivel, unidad = "Directorio", directorio.directorio_nombre
                    else:
                        _registrar_aviso(
                            avisos, fila_excel, dni, [COL_CUOF_DEPARTAMENTO, COL_DEPARTAMENTO_NOMBRE],
                            f"DNI {dni}: no existe ninguna unidad organizativa con C.U.O.F. {cuof} "
                            f"(xlsx: '{nombre_xlsx}'), se salteo oficina.",
                        )
                        return None

        if oficina is None:
            _registrar_aviso(
                avisos, fila_excel, dni, [COL_CUOF_DEPARTAMENTO, COL_DEPARTAMENTO_NOMBRE],
                f"DNI {dni}: existe {nivel} '{unidad}' (C.U.O.F. {cuof}) pero no tiene una Oficina "
                f"asociada; correr 'crear_oficinas' y reintentar. Se salteo oficina.",
            )
        return oficina
