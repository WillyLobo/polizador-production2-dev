import logging
import sqlite3

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.validators import get_cuil
from personalizador.management.commands.importar_informe_ipduv import (
    COL_DNI,
    DEFAULT_PATH,
    FIRST_DATA_ROW,
    SHEET_NAME,
    _blank,
    _normalize_text,
    _parse_dni,
)
from personalizador.models import Agente, GeneroAgente

logger = logging.getLogger(__name__)

DEFAULT_PADRON_DB = "/home/willy/dev/padron/padron/db.sqlite3"

GENERO_POR_TX_GENERO = {
    "M": "Masculino",
    "F": "Femenino",
}


class Command(BaseCommand):
    help = (
        "Crea los Agente que figuran en la hoja 'PLANTA PERMANENTE' de "
        "informe_ipduv.xlsx (por D.N.I.) pero todavia no existen en la base, "
        "buscando sus datos (apellido, nombre, sexo, domicilio) en el padron "
        "electoral (listado_padron.NU_MATRICULA = dni) y calculando el CUIL "
        "con core.validators.get_cuil(). No pisa Agentes existentes: solo crea "
        "los que faltan. Los D.N.I. que no aparecen en el padron, o cuyo "
        "TX_GENERO no sea M/F, se saltean con un aviso. Correr "
        "importar_informe_ipduv despues de este comando para completar los "
        "demas campos (categoria, oficina, etc.) de los Agentes recien creados."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", default=str(DEFAULT_PATH),
            help="Ruta al xlsx (default: %(default)s)",
        )
        parser.add_argument(
            "--padron-db", default=DEFAULT_PADRON_DB,
            help="Ruta a la base sqlite del padron (default: %(default)s)",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="No guarda cambios en la base de datos, solo informa.",
        )

    def handle(self, *args, **options):
        import openpyxl

        path = options["file"]
        padron_path = options["padron_db"]
        dry_run = options["dry_run"]
        verbosity = options.get("verbosity", 1)

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontro el archivo: {path}")

        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"La hoja '{SHEET_NAME}' no existe en {path}.")
        ws = wb[SHEET_NAME]

        try:
            padron_con = sqlite3.connect(padron_path)
        except sqlite3.Error as exc:
            raise CommandError(f"No se pudo abrir la base del padron ({padron_path}): {exc}")

        dnis_faltantes = []
        vistos = set()
        for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
            if _blank(row[COL_DNI]):
                continue
            dni = _parse_dni(row[COL_DNI])
            if dni is None or dni in vistos:
                continue
            vistos.add(dni)
            if not Agente.objects.filter(dni=dni).exists():
                dnis_faltantes.append(dni)

        stats = {"creados": 0, "no_en_padron": 0, "genero_no_soportado": 0, "genero_no_cargado": 0}
        avisos = []

        with transaction.atomic():
            for dni in dnis_faltantes:
                cur = padron_con.execute(
                    "SELECT TX_APELLIDO, TX_NOMBRE, TX_GENERO, TX_DOMICILIO "
                    "FROM listado_padron WHERE NU_MATRICULA = ?",
                    (str(dni),),
                )
                fila = cur.fetchone()
                if not fila:
                    stats["no_en_padron"] += 1
                    avisos.append(f"DNI {dni}: no encontrado en el padron, no se creo el Agente.")
                    continue

                apellido, nombre, tx_genero, domicilio = fila
                genero_nombre = GENERO_POR_TX_GENERO.get((tx_genero or "").strip().upper())
                if genero_nombre is None:
                    stats["genero_no_soportado"] += 1
                    avisos.append(
                        f"DNI {dni}: TX_GENERO del padron ('{tx_genero}') no es M/F, no se puede "
                        f"calcular CUIL de forma confiable; no se creo el Agente."
                    )
                    continue

                genero = GeneroAgente.objects.filter(generoagente_nombre=genero_nombre).first()
                if genero is None:
                    stats["genero_no_cargado"] += 1
                    avisos.append(
                        f"DNI {dni}: no existe GeneroAgente '{genero_nombre}' en la base "
                        f"(correr fill_data/fill_fields); no se creo el Agente."
                    )
                    continue

                cuil_crudo = get_cuil(str(dni), tx_genero)
                cuil = f"{cuil_crudo[:2]}-{cuil_crudo[2:10]}-{cuil_crudo[10:]}"

                stats["creados"] += 1
                if not dry_run:
                    Agente.objects.create(
                        dni=dni,
                        agente_nombres=_normalize_text(nombre).title(),
                        agente_apellidos=_normalize_text(apellido).title(),
                        sexo=genero,
                        cuil=cuil,
                        domicilio_direccion=_normalize_text(domicilio) if not _blank(domicilio) else None,
                        agente_verificado_contra_padron=True,
                    )
                if verbosity >= 2:
                    self.stdout.write(f"DNI {dni}: creado ({_normalize_text(apellido).title()}, {_normalize_text(nombre).title()}) CUIL {cuil}")

            if dry_run:
                transaction.set_rollback(True)

        padron_con.close()

        if verbosity >= 1:
            self.stdout.write("")
            if dry_run:
                self.stdout.write(self.style.NOTICE("Modo dry-run: no se guardo ningun cambio en la base de datos."))
            self.stdout.write(self.style.MIGRATE_HEADING("Resumen:"))
            self.stdout.write(self.style.SUCCESS(f"    Agentes creados: {stats['creados']}"))
            self.stdout.write(self.style.WARNING(f"    No encontrados en el padron: {stats['no_en_padron']}"))
            self.stdout.write(self.style.WARNING(f"    TX_GENERO no soportado (no M/F): {stats['genero_no_soportado']}"))
            self.stdout.write(self.style.WARNING(f"    GeneroAgente faltante en la base: {stats['genero_no_cargado']}"))

        if avisos and verbosity >= 1:
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(f"Avisos ({len(avisos)}):"))
            for aviso in avisos:
                self.stdout.write(f"    {self.style.WARNING(aviso)}")
