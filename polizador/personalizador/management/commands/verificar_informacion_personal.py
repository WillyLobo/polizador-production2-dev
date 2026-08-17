import json
import logging
import sqlite3
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from core.validators import get_cuil
from personalizador.management.commands.importar_informe_ipduv import (
    _blank,
    _normalize_text,
    _parse_dni,
)
from personalizador.models import Agente

logger = logging.getLogger(__name__)

DEFAULT_PATH = settings.BASE_DIR.parent / "env" / ".snipets" / "Informacion_personal.xlsx"
DEFAULT_PADRON_DB = "/home/willy/dev/padron/padron/db.sqlite3"
DEFAULT_OUTPUT = settings.BASE_DIR.parent / "env" / ".snipets" / "agentes_informacion_personal_faltantes.json"

SHEET_NAME = "Información personal"

# Columnas de la hoja (0-indexadas). El archivo trae 10 filas de reglas antes
# del header real (fila 11), asi que los datos empiezan en la fila 12.
COL_ID = 0
COL_NOMBRE = 1
COL_APELLIDO = 2

FIRST_DATA_ROW = 12

GENERO_POR_TX_GENERO = {
    "M": "Masculino",
    "F": "Femenino",
}


class Command(BaseCommand):
    help = (
        "Compara los agentes de Informacion_personal.xlsx (columna ID) contra "
        "personalizador.Agente (campo dni). Los IDs que no tienen un Agente "
        "cargado en la base se buscan en el padron electoral (listado_padron."
        "NU_MATRICULA = dni), igual que dni_check, y se exportan a un JSON "
        "compatible con importar_agentes_padron para aplicarse en produccion. "
        "Los IDs que no aparecen en el padron, o cuyo TX_GENERO no sea M/F, se "
        "informan como aviso y no se incluyen en el JSON."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", default=str(DEFAULT_PATH),
            help="Ruta al xlsx (default: %(default)s)",
        )
        parser.add_argument(
            "--padron-db", default=DEFAULT_PADRON_DB,
            help="Ruta a la base sqlite del padron (default: %(default)s)",
        )
        parser.add_argument(
            "--output", default=str(DEFAULT_OUTPUT),
            help="Ruta del JSON a generar (default: %(default)s)",
        )

    def handle(self, *args, **options):
        import openpyxl

        path = options["file"]
        padron_path = options["padron_db"]
        output_path = Path(options["output"])
        verbosity = options.get("verbosity", 1)

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontro el archivo: {path}")

        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        filas_excel = {}
        duplicados = []
        for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
            if _blank(row[COL_ID]):
                continue
            dni = _parse_dni(row[COL_ID])
            if dni is None:
                continue
            if dni in filas_excel:
                duplicados.append(dni)
                continue
            nombre = _normalize_text(row[COL_NOMBRE]) if not _blank(row[COL_NOMBRE]) else ""
            apellido = _normalize_text(row[COL_APELLIDO]) if not _blank(row[COL_APELLIDO]) else ""
            filas_excel[dni] = f"{apellido}, {nombre}".strip(", ")

        dnis_db = {int(dni) for dni in Agente.objects.values_list("dni", flat=True)}

        faltantes = {dni: nombre for dni, nombre in filas_excel.items() if dni not in dnis_db}

        if verbosity >= 1:
            self.stdout.write(self.style.MIGRATE_HEADING("Resumen:"))
            self.stdout.write(f"    Agentes en el excel: {len(filas_excel)}")
            self.stdout.write(f"    Agentes en la base (Agente.dni): {len(dnis_db)}")
            self.stdout.write(self.style.WARNING(f"    IDs del excel sin Agente cargado en la base: {len(faltantes)}"))
            if duplicados:
                self.stdout.write(self.style.WARNING(f"    IDs duplicados en el excel (se ignoro la repeticion): {len(duplicados)}"))

        registros, avisos = self._buscar_en_padron(faltantes, padron_path)

        for registro in registros:
            nombreyapellido_excel = filas_excel.get(registro["dni"], "")
            nombreyapellido_padron = f"{registro['agente_apellidos']}, {registro['agente_nombres']}"
            if nombreyapellido_excel and nombreyapellido_excel.upper() != nombreyapellido_padron.upper():
                avisos.append(
                    f"DNI {registro['dni']}: el nombre del excel ({nombreyapellido_excel}) no "
                    f"coincide con el del padron ({nombreyapellido_padron}); se exporto con el "
                    f"dato del padron, revisar a mano."
                )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8") as f:
            json.dump(registros, f, ensure_ascii=False, indent=2)

        if verbosity >= 1:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS(f"Exportados {len(registros)} agentes a {output_path}"))

        if avisos and verbosity >= 1:
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(f"Avisos ({len(avisos)}):"))
            for aviso in avisos:
                self.stdout.write(f"    {self.style.WARNING(aviso)}")

    def _buscar_en_padron(self, faltantes, padron_path):
        try:
            padron_con = sqlite3.connect(padron_path)
        except sqlite3.Error as exc:
            raise CommandError(f"No se pudo abrir la base del padron ({padron_path}): {exc}")

        registros = []
        avisos = []
        for dni in sorted(faltantes):
            cur = padron_con.execute(
                "SELECT TX_APELLIDO, TX_NOMBRE, TX_GENERO, TX_DOMICILIO "
                "FROM listado_padron WHERE NU_MATRICULA = ?",
                (str(dni),),
            )
            fila = cur.fetchone()
            if not fila:
                avisos.append(f"DNI {dni}: no encontrado en el padron, no se incluyo en el JSON.")
                continue

            apellido, nombre, tx_genero, domicilio = fila
            genero_nombre = GENERO_POR_TX_GENERO.get((tx_genero or "").strip().upper())
            if genero_nombre is None:
                avisos.append(
                    f"DNI {dni}: TX_GENERO del padron ('{tx_genero}') no es M/F, no se puede "
                    f"calcular CUIL de forma confiable; no se incluyo en el JSON."
                )
                continue

            cuil_crudo = get_cuil(str(dni), tx_genero)
            cuil = f"{cuil_crudo[:2]}-{cuil_crudo[2:10]}-{cuil_crudo[10:]}"

            registro = {
                "dni": dni,
                "agente_nombres": _normalize_text(nombre).title(),
                "agente_apellidos": _normalize_text(apellido).title(),
                "cuil": cuil,
                "sexo": genero_nombre,
                "agente_verificado_contra_padron": True,
            }
            if not _blank(domicilio):
                registro["domicilio_direccion"] = _normalize_text(domicilio).title()

            registros.append(registro)

        padron_con.close()
        return registros, avisos
